from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
import random, string, bcrypt
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (Member, Admin, Seat, Membership, Payment,
                     Attendance, SeatChangeRequest, Notification,
                     Setting, EmailOtp)
from .serializers import (MemberSerializer, SeatSerializer,
                           SeatDetailSerializer, MembershipSerializer,
                           PaymentSerializer, AttendanceSerializer,
                           SeatChangeRequestSerializer,
                           NotificationSerializer, SettingSerializer)


# ─── HELPERS ─────────────────────────────────────────────────
def hash_password(password):
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def check_password(password, hashed):
    try:
        if isinstance(hashed, str):
            hashed = hashed.encode('utf-8')
        if isinstance(password, str):
            password = password.encode('utf-8')
        return bcrypt.checkpw(password, hashed)
    except Exception as e:
        print(f"Password check error: {e}")
        return False

def generate_otp():
    return ''.join(random.choices(string.digits, k=6))

def create_notification(member, title, message, notif_type='general'):
    Notification.objects.create(
        member=member, title=title,
        message=message, type=notif_type
    )

def get_monthly_fee():
    try:
        return float(Setting.objects.get(key='monthly_fee').value)
    except:
        return 500.0


# ─── REGISTRATION & OTP ──────────────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
def send_otp(request):
    """Send OTP to email before registration"""
    email = request.data.get('email', '').strip().lower()
    if not email:
        return Response({'error': 'Email is required.'},
                        status=status.HTTP_400_BAD_REQUEST)
    if Member.objects.filter(email=email).exists():
        return Response({'error': 'Email already registered.'},
                        status=status.HTTP_400_BAD_REQUEST)

    # Invalidate old OTPs
    EmailOtp.objects.filter(email=email, is_used=False).update(is_used=True)

    otp_code   = generate_otp()
    expires_at = timezone.now() + timezone.timedelta(minutes=10)
    EmailOtp.objects.create(email=email, otp_code=otp_code,
                             expires_at=expires_at)

    try:
        send_mail(
            subject='Your Library Registration OTP',
            message=f'Your OTP is: {otp_code}\nValid for 10 minutes.',
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
        )
    except Exception as e:
        return Response({'error': f'Failed to send email: {str(e)}'},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    return Response({'message': 'OTP sent successfully.'})


@api_view(['POST'])
@permission_classes([AllowAny])
def verify_otp(request):
    """Verify OTP"""
    email    = request.data.get('email', '').strip().lower()
    otp_code = request.data.get('otp', '').strip()

    otp = EmailOtp.objects.filter(
        email=email, otp_code=otp_code,
        is_used=False, expires_at__gt=timezone.now()
    ).first()

    if not otp:
        return Response({'error': 'Invalid or expired OTP.'},
                        status=status.HTTP_400_BAD_REQUEST)

    otp.is_used = True
    otp.save()
    return Response({'message': 'OTP verified successfully.'})


@api_view(['POST'])
@permission_classes([AllowAny])
def register_member(request):
    """Register new member after OTP verification"""
    data = request.data

    required = ['full_name', 'father_name', 'mobile', 'email',
                'password', 'address']
    for field in required:
        if not data.get(field):
            return Response({'error': f'{field} is required.'},
                            status=status.HTTP_400_BAD_REQUEST)

    if Member.objects.filter(mobile=data['mobile']).exists():
        return Response({'error': 'Mobile already registered.'},
                        status=status.HTTP_400_BAD_REQUEST)
    if Member.objects.filter(email=data['email']).exists():
        return Response({'error': 'Email already registered.'},
                        status=status.HTTP_400_BAD_REQUEST)

    member = Member.objects.create(
        full_name         = data['full_name'],
        father_name       = data['father_name'],
        mobile            = data['mobile'],
        email             = data['email'].lower(),
        password_hash     = hash_password(data['password']),
        address           = data['address'],
        profile_photo     = request.FILES.get('profile_photo'),
        aadhar_photo      = request.FILES.get('aadhar_photo'),
        is_email_verified = True,
    )

    return Response({
        'message': 'Registration successful!',
        'member_id': member.member_id
    }, status=status.HTTP_201_CREATED)


# ─── LOGIN ───────────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
def member_login(request):
    """Member login with mobile + password"""
    mobile   = request.data.get('mobile', '').strip()
    password = request.data.get('password', '').strip()

    try:
        member = Member.objects.get(mobile=mobile)
    except Member.DoesNotExist:
        return Response({'error': 'Invalid mobile or password.'},
                        status=status.HTTP_401_UNAUTHORIZED)

    if not member.is_active:
        return Response({'error': 'Account is inactive. Contact admin.'},
                        status=status.HTTP_403_FORBIDDEN)

    if not check_password(password, member.password_hash):
        return Response({'error': 'Invalid mobile or password.'},
                        status=status.HTTP_401_UNAUTHORIZED)

    # Get seat info
    seat_number = None
    try:
        seat_number = member.seat.seat_number
    except:
        pass

    # Get active membership
    membership = member.memberships.filter(status='active').first()

    return Response({
        'message':    'Login successful.',
        'member_id':  member.member_id,
        'full_name':  member.full_name,
        'seat_number': seat_number,
        'membership': MembershipSerializer(membership).data if membership else None,
    })


@api_view(['POST'])
@permission_classes([AllowAny])
def admin_login(request):
    """Admin login"""
    mobile   = request.data.get('mobile', '').strip()
    password = request.data.get('password', '').strip()

    try:
        admin = Admin.objects.get(mobile=mobile)
    except Admin.DoesNotExist:
        return Response({'error': 'Invalid credentials.'},
                        status=status.HTTP_401_UNAUTHORIZED)

    if not check_password(password, admin.password_hash):
        return Response({'error': 'Invalid credentials.'},
                        status=status.HTTP_401_UNAUTHORIZED)

    return Response({
        'message':  'Admin login successful.',
        'admin_id': admin.admin_id,
        'name':     admin.name,
    })


# ─── SEATS ───────────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([AllowAny])
def seat_map(request):
    """All 90 seats with occupancy status"""
    seats = Seat.objects.select_related('member').order_by('seat_number')
    return Response(SeatSerializer(seats, many=True).data)


@api_view(['GET'])
@permission_classes([AllowAny])
def seat_detail(request, seat_id):
    try:
        seat = Seat.objects.select_related('member').get(seat_id=seat_id)
    except Seat.DoesNotExist:
        return Response({'error': 'Seat not found.'},
                        status=status.HTTP_404_NOT_FOUND)
    try:
        return Response(SeatDetailSerializer(seat).data)
    except Exception as e:
        return Response({'error': str(e)},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@permission_classes([AllowAny])
def assign_seat(request):
    """Admin assigns seat to member"""
    member_id = request.data.get('member_id')
    seat_id   = request.data.get('seat_id')

    try:
        member = Member.objects.get(member_id=member_id)
        seat   = Seat.objects.get(seat_id=seat_id)
    except (Member.DoesNotExist, Seat.DoesNotExist):
        return Response({'error': 'Member or seat not found.'},
                        status=status.HTTP_404_NOT_FOUND)

    if seat.is_occupied:
        return Response({'error': 'Seat is already occupied.'},
                        status=status.HTTP_400_BAD_REQUEST)

    # Free old seat if any
    try:
        old_seat = member.seat
        old_seat.member      = None
        old_seat.is_occupied = False
        old_seat.assigned_date = None
        old_seat.save()
    except:
        pass

    seat.member        = member
    seat.is_occupied   = True
    seat.assigned_date = timezone.now().date()
    seat.save()

    create_notification(member, 'Seat Assigned',
                        f'You have been assigned Seat {seat.seat_number}.', 'seat')

    return Response({'message': f'Seat {seat.seat_number} assigned successfully.'})


# ─── SEAT CHANGE REQUESTS ────────────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
def request_seat_change(request):
    """Member requests a different seat"""
    member_id          = request.data.get('member_id')
    requested_seat_id  = request.data.get('requested_seat_id')

    try:
        member         = Member.objects.get(member_id=member_id)
        current_seat   = member.seat
        requested_seat = Seat.objects.get(seat_id=requested_seat_id)
    except Member.DoesNotExist:
        return Response({'error': 'Member not found.'},
                        status=status.HTTP_404_NOT_FOUND)
    except Seat.DoesNotExist:
        return Response({'error': 'Requested seat not found.'},
                        status=status.HTTP_404_NOT_FOUND)
    except Exception:
        return Response({'error': 'You do not have an assigned seat.'},
                        status=status.HTTP_400_BAD_REQUEST)

    if requested_seat.is_occupied:
        return Response({'error': 'Requested seat is already occupied.'},
                        status=status.HTTP_400_BAD_REQUEST)

    # Check no pending request exists
    if SeatChangeRequest.objects.filter(
            member=member, status='pending').exists():
        return Response({'error': 'You already have a pending seat change request.'},
                        status=status.HTTP_400_BAD_REQUEST)

    SeatChangeRequest.objects.create(
        member=member,
        current_seat=current_seat,
        requested_seat=requested_seat,
    )

    return Response({'message': 'Seat change request submitted. Waiting for admin approval.'})


@api_view(['GET'])
@permission_classes([AllowAny])
def pending_seat_requests(request):
    """Admin views all pending seat change requests"""
    requests = SeatChangeRequest.objects.filter(
        status='pending').select_related(
        'member', 'current_seat', 'requested_seat')
    return Response(SeatChangeRequestSerializer(requests, many=True).data)


@api_view(['POST'])
@permission_classes([AllowAny])
def resolve_seat_request(request, request_id):
    """Admin approves or rejects seat change request"""
    action   = request.data.get('action')   # 'approve' or 'reject'
    admin_id = request.data.get('admin_id')

    try:
        scr   = SeatChangeRequest.objects.get(request_id=request_id)
        admin = Admin.objects.get(admin_id=admin_id)
    except (SeatChangeRequest.DoesNotExist, Admin.DoesNotExist):
        return Response({'error': 'Request or admin not found.'},
                        status=status.HTTP_404_NOT_FOUND)

    if scr.status != 'pending':
        return Response({'error': 'Request already resolved.'},
                        status=status.HTTP_400_BAD_REQUEST)

    if action == 'approve':
        if scr.requested_seat.is_occupied:
            return Response({'error': 'Seat was occupied in the meantime.'},
                            status=status.HTTP_400_BAD_REQUEST)

        # Free current seat
        scr.current_seat.member      = None
        scr.current_seat.is_occupied = False
        scr.current_seat.assigned_date = None
        scr.current_seat.save()

        # Assign new seat
        scr.requested_seat.member      = scr.member
        scr.requested_seat.is_occupied = True
        scr.requested_seat.assigned_date = timezone.now().date()
        scr.requested_seat.save()

        scr.status      = 'approved'
        msg = f'Your seat change to Seat {scr.requested_seat.seat_number} has been approved!'
        create_notification(scr.member, 'Seat Change Approved', msg, 'seat')

    else:
        scr.status = 'rejected'
        msg = f'Your seat change request to Seat {scr.requested_seat.seat_number} was rejected.'
        create_notification(scr.member, 'Seat Change Rejected', msg, 'seat')

    scr.resolved_at  = timezone.now()
    scr.resolved_by  = admin
    scr.save()

    return Response({'message': f'Request {scr.status} successfully.'})


# ─── MEMBERSHIP ──────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
def add_membership(request):
    """Admin adds/renews membership for a member"""
    member_id = request.data.get('member_id')
    admin_id  = request.data.get('admin_id')
    start_str = request.data.get('start_date')

    try:
        member = Member.objects.get(member_id=member_id)
        admin  = Admin.objects.get(admin_id=admin_id)
    except (Member.DoesNotExist, Admin.DoesNotExist):
        return Response({'error': 'Member or admin not found.'},
                        status=status.HTTP_404_NOT_FOUND)

    # Expire any current active membership
    member.memberships.filter(status='active').update(status='expired')

    from datetime import date, timedelta
    start_date = date.fromisoformat(start_str) if start_str else date.today()
    end_date   = start_date.replace(day=1)
    # End date = last day of month
    if start_date.month == 12:
        end_date = date(start_date.year + 1, 1, start_date.day)
    else:
        end_date = start_date + timedelta(days=30)

    fee = get_monthly_fee()

    membership = Membership.objects.create(
        member=member, start_date=start_date,
        end_date=end_date, fee_paid=fee, status='active'
    )

    Payment.objects.create(
        member=member, membership=membership,
        amount=fee, payment_mode='cash',
        recorded_by=admin
    )

    create_notification(
        member, 'Membership Renewed',
        f'Your membership is active from {start_date} to {end_date}. Fee: ₹{fee}',
        'payment'
    )

    return Response({
        'message':    'Membership added successfully.',
        'membership': MembershipSerializer(membership).data
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([AllowAny])
def cancel_membership(request, member_id):
    """Admin cancels membership — frees seat"""
    try:
        member = Member.objects.get(member_id=member_id)
    except Member.DoesNotExist:
        return Response({'error': 'Member not found.'},
                        status=status.HTTP_404_NOT_FOUND)

    member.memberships.filter(status='active').update(status='cancelled')

    # Free seat
    try:
        seat = member.seat
        seat.member      = None
        seat.is_occupied = False
        seat.assigned_date = None
        seat.save()
    except:
        pass

    member.is_active = False
    member.save()

    create_notification(member, 'Membership Cancelled',
                        'Your membership has been cancelled.', 'general')

    return Response({'message': 'Membership cancelled and seat freed.'})


# ─── ATTENDANCE ──────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
def mark_attendance(request):
    """Admin marks a member present"""
    member_id = request.data.get('member_id')
    admin_id  = request.data.get('admin_id')
    date_str  = request.data.get('date')

    try:
        member = Member.objects.get(member_id=member_id)
        admin  = Admin.objects.get(admin_id=admin_id)
    except (Member.DoesNotExist, Admin.DoesNotExist):
        return Response({'error': 'Member or admin not found.'},
                        status=status.HTTP_404_NOT_FOUND)

    from datetime import date
    attendance_date = date.fromisoformat(date_str) if date_str else date.today()

    attendance, created = Attendance.objects.get_or_create(
        member=member, date=attendance_date,
        defaults={'marked_by': admin}
    )

    if not created:
        return Response({'message': 'Attendance already marked for today.'})

    return Response({'message': f'Attendance marked for {member.full_name}.'})


@api_view(['GET'])
@permission_classes([AllowAny])
def today_attendance(request):
    """Admin sees today's attendance"""
    today = timezone.now().date()
    records = Attendance.objects.filter(
        date=today).select_related('member')
    return Response({
        'date':          str(today),
        'total_present': records.count(),
        'members':       AttendanceSerializer(records, many=True).data
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def member_attendance(request, member_id):
    """Member views own attendance history"""
    try:
        member = Member.objects.get(member_id=member_id)
    except Member.DoesNotExist:
        return Response({'error': 'Member not found.'},
                        status=status.HTTP_404_NOT_FOUND)

    today = timezone.now().date()
    records = member.attendance.filter(
        date__year=today.year,
        date__month=today.month
    ).order_by('-date')

    return Response({
        'member_name':   member.full_name,
        'month':         today.strftime('%B %Y'),
        'total_present': records.count(),
        'records':       AttendanceSerializer(records, many=True).data
    })


# ─── PAYMENTS ─────────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([AllowAny])
def pending_payments(request):
    """Admin sees all members with no active membership"""
    all_members = Member.objects.filter(is_active=True)
    unpaid = [m for m in all_members
              if not m.memberships.filter(status='active').exists()]
    return Response({
        'total_unpaid': len(unpaid),
        'members': MemberSerializer(unpaid, many=True).data
    })


@api_view(['GET'])
@permission_classes([AllowAny])
def member_payments(request, member_id):
    try:
        member = Member.objects.get(member_id=member_id)
    except Member.DoesNotExist:
        return Response({'error': 'Member not found.'},
                        status=status.HTTP_404_NOT_FOUND)
    try:
        payments = Payment.objects.filter(member=member).order_by('-payment_date')
        return Response(PaymentSerializer(payments, many=True).data)
    except Exception as e:
        return Response({'error': str(e)},
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ─── NOTIFICATIONS ────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([AllowAny])
def member_notifications(request, member_id):
    """Member views notifications"""
    try:
        member = Member.objects.get(member_id=member_id)
    except Member.DoesNotExist:
        return Response({'error': 'Member not found.'},
                        status=status.HTTP_404_NOT_FOUND)

    notifs = member.notifications.order_by('-created_at')
    return Response(NotificationSerializer(notifs, many=True).data)


@api_view(['POST'])
@permission_classes([AllowAny])
def mark_notification_read(request, notification_id):
    """Mark a notification as read"""
    try:
        notif = Notification.objects.get(notification_id=notification_id)
        notif.is_read = True
        notif.save()
        return Response({'message': 'Marked as read.'})
    except Notification.DoesNotExist:
        return Response({'error': 'Notification not found.'},
                        status=status.HTTP_404_NOT_FOUND)


# ─── MEMBERS (Admin) ──────────────────────────────────────────
@api_view(['GET'])
@permission_classes([AllowAny])
def all_members(request):
    """Admin views all members"""
    search = request.query_params.get('search', '')
    members = Member.objects.filter(is_active=True)
    if search:
        members = members.filter(
            full_name__icontains=search
        ) | members.filter(
            mobile__icontains=search
        )
    return Response(MemberSerializer(members, many=True).data)


@api_view(['GET'])
@permission_classes([AllowAny])
def member_detail(request, member_id):
    """Get single member details"""
    try:
        member = Member.objects.get(member_id=member_id)
        return Response(MemberSerializer(member).data)
    except Member.DoesNotExist:
        return Response({'error': 'Member not found.'},
                        status=status.HTTP_404_NOT_FOUND)


# ─── SETTINGS ─────────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([AllowAny])
def get_settings(request):
    """Get all settings"""
    s = Setting.objects.all()
    return Response(SettingSerializer(s, many=True).data)


@api_view(['POST'])
@permission_classes([AllowAny])
def update_setting(request):
    """Admin updates a setting e.g. monthly fee"""
    key   = request.data.get('key')
    value = request.data.get('value')
    if not key or value is None:
        return Response({'error': 'Key and value required.'},
                        status=status.HTTP_400_BAD_REQUEST)
    setting, _ = Setting.objects.get_or_create(key=key)
    setting.value = str(value)
    setting.save()
    return Response({'message': f'Setting {key} updated to {value}.'})


@api_view(['POST'])
@permission_classes([AllowAny])
def update_admin(request):
    """Admin updates their own profile"""
    admin_id  = request.data.get('admin_id')
    name      = request.data.get('name')
    mobile    = request.data.get('mobile')
    password  = request.data.get('password')

    try:
        admin = Admin.objects.get(admin_id=admin_id)
    except Admin.DoesNotExist:
        return Response({'error': 'Admin not found.'},
                        status=status.HTTP_404_NOT_FOUND)

    if name:   admin.name   = name
    if mobile: admin.mobile = mobile
    if password:
        admin.password_hash = hash_password(password)
    admin.save()

    return Response({'message': 'Profile updated successfully.'})

# ─── ANALYTICS ────────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([AllowAny])
def analytics_overview(request):
    """Main analytics data for Tableau / dashboard"""
    from django.db.models import Count, Sum
    from datetime import date

    today = date.today()

    # Seat occupancy
    total_seats    = Seat.objects.count()
    occupied_seats = Seat.objects.filter(is_occupied=True).count()
    vacant_seats   = total_seats - occupied_seats

    # Members
    total_members  = Member.objects.filter(is_active=True).count()
    active_members = Member.objects.filter(
        is_active=True,
        memberships__status='active'
    ).distinct().count()

    # Today attendance
    today_present = Attendance.objects.filter(date=today).count()

    # Monthly revenue
    current_month_revenue = Payment.objects.filter(
        payment_date__year=today.year,
        payment_date__month=today.month
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Expiring soon (next 3 days)
    from datetime import timedelta
    expiring_soon = Membership.objects.filter(
        status='active',
        end_date__lte=today + timedelta(days=3),
        end_date__gte=today
    ).count()

    # Pending payments
    all_active = Member.objects.filter(is_active=True)
    pending_count = sum(
        1 for m in all_active
        if not m.memberships.filter(status='active').exists()
    )

    return Response({
        'seats': {
            'total':    total_seats,
            'occupied': occupied_seats,
            'vacant':   vacant_seats,
            'occupancy_rate': round((occupied_seats / total_seats) * 100, 1)
        },
        'members': {
            'total':   total_members,
            'active':  active_members,
            'pending_payment': pending_count,
        },
        'today': {
            'date':    str(today),
            'present': today_present,
        },
        'finance': {
            'monthly_fee':            get_monthly_fee(),
            'current_month_revenue':  float(current_month_revenue),
        },
        'alerts': {
            'expiring_soon': expiring_soon,
        }
    })

# ─── EXCEL EXPORTS ───────────────────────────────────────────

@api_view(['GET'])
@permission_classes([AllowAny])
def export_members_excel(request):
    """Export all members to Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")

    headers = ['#', 'Full Name', 'Father Name', 'Mobile',
               'Email', 'Address', 'Status', 'Joined Date',
               'Seat Number', 'Membership Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64+col)].width = 20

    # Data
    members = Member.objects.filter(is_active=True).order_by('member_id')
    for row, member in enumerate(members, 2):
        # Get seat
        seat_number = '—'
        try:
            seat_number = member.seat.seat_number
        except: pass

        # Get membership
        membership = member.memberships.filter(status='active').first()
        membership_status = membership.status if membership else 'No membership'

        ws.cell(row=row, column=1,  value=row-1)
        ws.cell(row=row, column=2,  value=member.full_name)
        ws.cell(row=row, column=3,  value=member.father_name)
        ws.cell(row=row, column=4,  value=member.mobile)
        ws.cell(row=row, column=5,  value=member.email)
        ws.cell(row=row, column=6,  value=member.address)
        ws.cell(row=row, column=7,  value='Active' if member.is_active else 'Inactive')
        ws.cell(row=row, column=8,  value=str(member.created_at.date()))
        ws.cell(row=row, column=9,  value=str(seat_number))
        ws.cell(row=row, column=10, value=membership_status)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=members.xlsx'
    response['Access-Control-Allow-Origin'] = '*'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([AllowAny])
def export_payments_excel(request):
    """Export all payments to Excel"""
    month = request.query_params.get('month')
    year  = request.query_params.get('year')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="16A34A")

    headers = ['#', 'Member Name', 'Mobile', 'Amount (₹)',
               'Payment Mode', 'Payment Date', 'Month', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64+col)].width = 20

    payments = Payment.objects.select_related(
        'member').order_by('-payment_date')
    if month and year:
        payments = payments.filter(
            payment_date__month=month,
            payment_date__year=year
        )

    total = 0
    for row, payment in enumerate(payments, 2):
        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=payment.member.full_name)
        ws.cell(row=row, column=3, value=payment.member.mobile)
        ws.cell(row=row, column=4, value=float(payment.amount))
        ws.cell(row=row, column=5, value=payment.payment_mode.upper())
        ws.cell(row=row, column=6, value=str(payment.payment_date.date()))
        ws.cell(row=row, column=7, value=payment.payment_date.strftime('%B %Y'))
        ws.cell(row=row, column=8, value=payment.notes or '—')
        total += float(payment.amount)

    # Total row
    total_row = payments.count() + 2
    ws.cell(row=total_row, column=3,
            value='TOTAL').font = Font(bold=True)
    ws.cell(row=total_row, column=4,
            value=total).font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=payments.xlsx'
    response['Access-Control-Allow-Origin'] = '*'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([AllowAny])
def export_attendance_excel(request):
    """Export attendance to Excel"""
    month = request.query_params.get('month', timezone.now().month)
    year  = request.query_params.get('year',  timezone.now().year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="9333EA")

    headers = ['#', 'Member Name', 'Mobile', 'Seat No',
               'Date', 'Day', 'Month']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64+col)].width = 20

    records = Attendance.objects.filter(
        date__month=month,
        date__year=year
    ).select_related('member').order_by('date', 'member__full_name')

    for row, record in enumerate(records, 2):
        seat_number = '—'
        try:
            seat_number = record.member.seat.seat_number
        except: pass

        ws.cell(row=row, column=1, value=row-1)
        ws.cell(row=row, column=2, value=record.member.full_name)
        ws.cell(row=row, column=3, value=record.member.mobile)
        ws.cell(row=row, column=4, value=str(seat_number))
        ws.cell(row=row, column=5, value=str(record.date))
        ws.cell(row=row, column=6, value=record.date.strftime('%A'))
        ws.cell(row=row, column=7, value=record.date.strftime('%B %Y'))

    # Summary
    summary_row = records.count() + 3
    ws.cell(row=summary_row, column=1,
            value=f'Total Records: {records.count()}').font = Font(bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=attendance.xlsx'
    response['Access-Control-Allow-Origin'] = '*'
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    wb.save(response)
    return response

@api_view(['POST'])
@permission_classes([AllowAny])
def update_total_seats(request):
    """Admin increases or decreases total seats"""
    new_total = int(request.data.get('total_seats', 90))
    current_total = Seat.objects.count()

    if new_total == current_total:
        return Response({'message': 'No change needed.'})

    if new_total > current_total:
        # Add new seats
        for i in range(current_total + 1, new_total + 1):
            Seat.objects.create(
                seat_number=i,
                is_occupied=False
            )
        # Update settings
        Setting.objects.filter(key='total_seats').update(value=str(new_total))
        return Response({
            'message': f'Added {new_total - current_total} seats. Total: {new_total}'
        })

    elif new_total < current_total:
        # Only remove vacant seats from the end
        seats_to_remove = Seat.objects.filter(
            is_occupied=False,
            seat_number__gt=new_total
        ).count()

        occupied_count = Seat.objects.filter(
            is_occupied=True,
            seat_number__gt=new_total
        ).count()

        if occupied_count > 0:
            return Response({
                'error': f'Cannot reduce — {occupied_count} occupied seats above seat #{new_total}. Please reassign them first.'
            }, status=status.HTTP_400_BAD_REQUEST)

        Seat.objects.filter(
            is_occupied=False,
            seat_number__gt=new_total
        ).delete()

        Setting.objects.filter(key='total_seats').update(value=str(new_total))
        return Response({
            'message': f'Removed {seats_to_remove} seats. Total: {new_total}'
        })